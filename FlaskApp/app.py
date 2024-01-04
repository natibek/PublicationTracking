from flask import Flask, render_template, redirect, url_for, request, send_file
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import PatternFill
from io import BytesIO
# from . import app

app = Flask(__name__)

bsd_faculty_cleaned_df = None
scopus_csv = None
output_excel = None
output_file_name = None
display_info = False

def is_faculty(authors: list, authors_and_affiliations: list): # find info as well if true and make dictionary
    """
    Checks if author is in the faculty list. Otherwise, checks if the affiliated department is in the list of bsd/som departments. 
    Returns dictionary containing name, department and track.
    
    Input: authors (list of strings) -> strings are last name first initial of identified uchicago faculty (department is not necessarily bsd/som)
           authors_and_affiliations (list of strings) -> author last name anf first initial and all the affiliation information including department, institution, location. Will use department 
    
    Output: author_info (dict) -> key is the author last first initial. Value is a dictionary with full name, department, and track keys.
    """
    departments = list(bsd_faculty_cleaned_df["Department Name"].unique()) # all the unique departments faculty can be part of
    
    author_info = {}

    for ind, author in enumerate(authors):
        is_found = (author.lower() in list(bsd_faculty_cleaned_df['Last Name First Initial'].str.lower()) or
                    author.lower() in list(bsd_faculty_cleaned_df['Previous Last Name First Initial'].str.lower()))     
        # condition for if the author is in the faculty listing

        if is_found: 
            # if in faculty, find the index in the faculty dataframe and extract name, department, and track
            last_name_found = (bsd_faculty_cleaned_df["Last Name First Initial"].str.lower() == author.lower())
            preferred_last_name_found = (bsd_faculty_cleaned_df["Previous Last Name First Initial"].str.lower() == author.lower())

            index = np.where(last_name_found | preferred_last_name_found)
            
    
            name = (bsd_faculty_cleaned_df.iloc[index[0][0]]['Name'])
            department = (bsd_faculty_cleaned_df.iloc[index[0][0]]['Department Name'])
            track = (bsd_faculty_cleaned_df.iloc[index[0][0]]['FACULTY TRACK'])

            author_info[author] = {"name": name, "department": department, "track": track}
        else:
            # if not in faculty, check if the affiliation is one of the deparments
            # the track is set to "-" due to ambiguity
            try:
                affiliation = authors_and_affiliations[ind][2]
                for department in departments:
                    if department in affiliation:
                        author_info[author] = {"name": author, "department": affiliation, "track": "-"}
                        break
            except:
                continue

    return author_info

def department_authors(publication: pd.DataFrame):
    """
    Parses through the "Authors and affiliations" column of the scopus file to find the authors and their affiliation to 
    uses the is_faculty to find all the authors in the deparment.

    Input: publication (pd.DataFrame) -> row of the scopus csv file

    Output: department_authors_info (dict) -> all the authors that are identified to be in the deparment. key is the author last first initial. 
                                              Value is a dictionary with full name, department, and track keys.
            all_authors (list of strings) -> last name first initial of all the authors in the order that they appear (both found in the department and not)
    """

    unprocessed = publication["Authors with affiliations"].split("; ")
    from_uchicago = [author for author in unprocessed if "University of Chicago" in author]
    authors_and_affiliations = [author_and_affiliation.split(", ") for author_and_affiliation in from_uchicago]
    authors = []
    formatting = "Normal"

    for author in authors_and_affiliations:#uses try in case there is odd formating
        try:
            authors.append(author[0] + " " + author[1][0])
        except:
            continue  

    all_processed = [author for author in unprocessed]
    all_authors_affilitions = [author_and_affiliation.split(", ") for author_and_affiliation in all_processed]

    all_authors = []
    for author in all_authors_affilitions:#uses try in case there is odd formating
        try:
            all_authors.append(author[0] + " " + author[1][0])
        except:
            continue

    no_affiliation = publication['Authors'].split(', ')
    names_no_affiliation = [author.split(" ") for author in no_affiliation]
    for name in names_no_affiliation:
        try:
            cur_name = name[0] + " " + name[-1][0]
            if cur_name not in all_authors:
                all_authors.append(cur_name)
                authors.append(cur_name)
                authors_and_affiliations.append(1)
                formatting = "Strange"
        except:
            continue

    department_authors_info = is_faculty(authors, authors_and_affiliations)
    return (department_authors_info, all_authors, formatting)

def get_corresponding_authors(publication: pd.DataFrame, department_authors_info: list, all_authors: list):
    '''
    Get the corresponding author for publication. Checks if the authors in the correspondence address are in the list of identified department affiliated authors. 
    If no correspondence address is provided, use the author in the authors list

    Input: publication (pd.DataFrame row)
           department_authors_info (list of dict) -> dict key is author last name first initial. Value is dictionary with department and track
           all_authors (list of strings) -> strings are last name first initial. All authors in order they appear

    Output: c_authors (tupble of string) -> (names, departments, tracks) aggregate of the information about corresponding authors
            c_authors_info (list of dict) -> dict key is changed to full name. 
    '''
    correspondence_address = publication['Correspondence Address']
    case = "Normal"

    if (type(correspondence_address) == str): # if there is a provided correspondence address, parse information to find the names of the corresponding authors
        addresses = correspondence_address.split('\n')
        names = [name.split(";")[0] for name in addresses]
        last_name_first_initial = [name.split(", ") for name in names]
        
        corresponding_authors = []
        for name in last_name_first_initial: #uses try in case there is odd formating
            try:
                corresponding_authors.append(name[0] + " " + name[1][0])
            except:
                continue
        
    else: # otherwise, use last author in list
        corresponding_authors = all_authors[-1]
        case = "Strange"
    
    c_authors_info = {c_author: department_authors_info[c_author] for c_author in corresponding_authors if c_author in department_authors_info}
    # dicationary has authors' last name first initial as the key
    # corresponding value is dictionary with full name, department, and track

    if len(c_authors_info) > 0: # if there is a corresponding author
        authors = []
        departments = []
        tracks = []

        for info in c_authors_info.values(): # aggregate the names, departments, and tracks 
            authors.append(info['name'])
            departments.append(info['department'])
            tracks.append(info['track'])

        c_authors = (", ".join(authors), ", ".join(departments), ", ".join(tracks))
        
        return c_authors, c_authors_info, case

    return ("-", "-", "-"), {}, case

def get_middle_authors(c_authors_info: list, department_authors_info: list):
    '''
    Gets all the authors that are not corresponding authors and in the department.

    Input: corresponding_authors (list of dict) -> strings are author Last Name First Initial
           department_authors_info (list of dict) -> dict key is author last name first initial. Value is dictionary with department and track

    Output: middle_authors (list of dict) -> dict key is changed to full name. 
    '''
    
    m_authors_info = {}
    for author in department_authors_info: # iterates over authors in the department to check if they are the corresponding author
        if author not in c_authors_info:
            m_authors_info[author] = department_authors_info[author]

    if len(m_authors_info) > 0: # in the case there is atleast one middle author
        authors = []
        departments = []
        tracks = []

        for info in m_authors_info.values(): # aggregate the names, departments, and tracks
            authors.append(info['name'])
            departments.append(info['department'])
            tracks.append(info['track'])

        m_authors = (", ".join(authors), ", ".join(departments), ", ".join(tracks))
        
        return m_authors
    return ("-", "-", "-") # default is dashes when no middle authors are found

def create_faculty_df(faculty_roaster_excel):
    bsd_faculty_df = pd.read_excel(faculty_roaster_excel)
    bsd_faculty_info = bsd_faculty_df[[" --FIRST NAME", " LAST NAME", " PREVIOUS LAST NAME"," DEPARTMENT NAME", "FACULTY TRACK"]]
    bsd_faculty_info.rename(columns={" --FIRST NAME": "First Name", " LAST NAME": "Last Name", " PREVIOUS LAST NAME": "Previous Last Name", " DEPARTMENT NAME": "Department Name"}, inplace=True)
    bsd_faculty_info = bsd_faculty_info.loc[bsd_faculty_info["FACULTY TRACK"].notna()]
    bsd_faculty_info["First Name"] = bsd_faculty_info["First Name"].apply(lambda x: x[0].upper() + x[1:])
    bsd_faculty_info["Last Name"] = bsd_faculty_info["Last Name"].apply(lambda x: x[0].upper() + x[1:])
    bsd_faculty_info["Name"] = bsd_faculty_info['First Name'] + " " + bsd_faculty_info["Last Name"]
    bsd_faculty_info["First Initial"] = bsd_faculty_info['First Name'].apply(lambda x: x[0])
    bsd_faculty_info["Last Name First Initial"] = bsd_faculty_info['Last Name'] + " " + bsd_faculty_info['First Initial']
    bsd_faculty_info["Previous Last Name First Initial"] = bsd_faculty_info['Previous Last Name'] + " " + bsd_faculty_info['First Initial']
    bsd_faculty_info["Department Name"] = bsd_faculty_info["Department Name"].str.replace('&', 'and')
    return bsd_faculty_info

@app.route("/")
def publication_tracking():
    global bsd_faculty_cleaned_df, scopus_csv, output_excel
    
    if bsd_faculty_cleaned_df is not None and scopus_csv is not None and output_excel is not None:
        all_valid = True
    else:
        all_valid = False

    if bsd_faculty_cleaned_df is None:
        faculty_status = False
    else:
        faculty_status = True

    if scopus_csv is None:
        scopus_status = False
    else:
        scopus_status = True

    if output_excel is None:
        output_status = False
    else:
        output_status = True

    if display_info:
        return render_template('info.html')
    else:
        return render_template('tracking.html', all_valid = all_valid, status = [faculty_status, scopus_status, output_status])

    
@app.route("/faculty_file", methods = ["POST"])
def faculty_file():
    input_faculty_file = request.files.get("faculty_roaster")
    
    temp_df = pd.read_excel(input_faculty_file)
    mandatory_cols = [" --FIRST NAME", " LAST NAME", " PREVIOUS LAST NAME"," DEPARTMENT NAME", "FACULTY TRACK"]
    missing_cols = [col for col in mandatory_cols if col not in temp_df.columns]
    if missing_cols == []:
        global bsd_faculty_cleaned_df
        bsd_faculty_cleaned_df = create_faculty_df(input_faculty_file)

    return redirect(url_for('publication_tracking'))

@app.route('/scopus_file', methods = ["POST"])
def scopus_file():
    input_scopus_file = request.files.get("scopus")

    temp_df = pd.read_csv(input_scopus_file)

    mandatory_cols = ['Authors', 'Title', 'Year', 'Source title', 'Link', 'Affiliations','Authors with affiliations', 'Correspondence Address']
    missing_cols = [col for col in mandatory_cols if col not in temp_df.columns]
    
    if missing_cols == []:
        global scopus_csv
        scopus_csv = temp_df

    return redirect(url_for('publication_tracking'))

@app.route('/output_file', methods = ["POST"])
def output_file():
    tracking_excel = request.files.get('tracking_excel')
    global output_excel, output_file_name
    output_file_name = tracking_excel.filename

    output_excel = BytesIO()
    tracking_excel.save(output_excel)
    output_excel.seek(0)
    
    return redirect(url_for('publication_tracking'))

@app.route('/submit')
def submit():

    global bsd_faculty_cleaned_df, scopus_csv, output_excel, output_file_name

    workbook = load_workbook(BytesIO(output_excel.read()))
    sheet = workbook.active

    for publication in range(len(scopus_csv)):
        cur_publication = scopus_csv.iloc[publication]
        
        department_authors_info, all_authors, formatting = department_authors(cur_publication)

        if len(department_authors_info) > 0:
            c_authors, c_authors_info, case = get_corresponding_authors(cur_publication, department_authors_info, all_authors)
            m_authors = get_middle_authors(c_authors_info, department_authors_info)
        
            c_authors_names, c_departments, c_tracks = c_authors
            m_authors_names, m_departments, m_tracks = m_authors

            title = cur_publication["Title"]    
            journal = cur_publication["Source title"]
            link = cur_publication["Link"]
            date = cur_publication['Year']
            
            output_data = [c_authors_names, c_departments, c_tracks, m_authors_names, m_departments, m_tracks, journal, title, date, link]
            sheet.append(output_data)
            
            if formatting == "Strange" and case == "Strange":
                current_row_index = sheet.max_row
                fill = PatternFill(start_color="0000CCFF", end_color="0000CCFF", fill_type="solid")

                for cell in sheet[current_row_index]:
                    cell.fill = fill
            elif case == "Strange":
                current_row_index = sheet.max_row
                fill = PatternFill(start_color="00FFFF99", end_color="00FFFF99", fill_type="solid")

                for cell in sheet[current_row_index]:
                    cell.fill = fill

            elif formatting == "Strange":
                current_row_index = sheet.max_row
                fill = PatternFill(start_color="00008000", end_color="00008000", fill_type="solid")

                for cell in sheet[current_row_index]:
                    cell.fill = fill

    modified_output = BytesIO()
    workbook.save(modified_output)
    modified_output.seek(0)
    download_file_name = "(" + str(datetime.now().month) + "_" + str(datetime.now().day) + "_" + str(datetime.now().year) + ") " + output_file_name 

    bsd_faculty_cleaned_df = scopus_csv =  output_excel = output_file_name = None
    
    return send_file(
        modified_output,
        as_attachment=True,
        download_name= download_file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/info", methods = ["GET"])
def info():
    global display_info
    display_info = not display_info
    
    return redirect(url_for('publication_tracking'))


# if __name__ == "__main__":
#     app.run(debug=True)
